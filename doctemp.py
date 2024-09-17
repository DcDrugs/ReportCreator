"""файл для работы с excel и docx
"""

from docxtpl import DocxTemplate
from openpyxl import load_workbook as l_workbook
from openpyxl.utils.cell import column_index_from_string
import re
from jinja2functions import jin, wrapper_error
from logger import logging
import magic2functions
from memoization import cached
import io

def load_workbook(filename : str, *args, **kwargs):
    """Функция для загрузки excel файла

    Args:
        filename (str): путь к файлу для его открытия

    Returns:
        Workbook: api для работы с excel документом
    """
    
    # обходим проблему с тем, что openpyxl не дает перезаписывать в excel файл, если приложение открыто
    in_mem_file = None
    with open(filename, "rb") as f:
        in_mem_file = io.BytesIO(f.read())
        
    return l_workbook(in_mem_file, *args, **kwargs)

# регулярка для поиска "{ ИмяЛиста }_{ Колонка }{ Строка }"
regex_excel_sheet = re.compile(r'(.+)_([A-Z]+)(\d*)')


class ExcelParseItem:
    """Класс для парсинга позиций в excel
    """
    sheet = None
    column = None
    row = None
    
    def __init__(self, sheet = None, col = None, row = None):
        """инициализация

        Args:
            sheet (str): Имя листа. Defaults to None.
            col (str): Колонка. Defaults to None.
            row (int): Строка. Defaults to None.
        """
        self.sheet = sheet
        self.column = col
        self.row = row
        
    def is_none(self):
        """Проверка на существование

        Returns:
            bool: являеться данный объект пустым
        """
        return self.sheet == None or self.column == None    
    
def try_parse(item: str, currentRow : int = None) -> ExcelParseItem:
    """функция для рабора строки на компоненты

    Args:
        item (str): строка пути в excel
        currentRow (int, optional): конкретная строка в excel документе. Defaults to None.

    Returns:
        ExcelParseItem: объект пути
    """
    result = regex_excel_sheet.search(item)
    if result == None:
        return ExcelParseItem()
    parsed_excel_sheets = result.groups()
    if len(parsed_excel_sheets) == 3:
        index = parsed_excel_sheets[2]
        if parsed_excel_sheets[2] == '' and currentRow != None:
            index = str(currentRow)
        return ExcelParseItem(parsed_excel_sheets[0], parsed_excel_sheets[1], index)
    else:
       return ExcelParseItem() 

def find_index(wb, excel_sheet_str: str, key: str, pos: int):
    """функция для поиска строки

    Args:
        wb (workbook): книга excel
        excel_sheet_str (str): тектовый путь, где необходимо искать ключ
        key (str): значение, которое необходимо искать
        pos (int): с какой строки искать

    Raises:
        KeyError: неверно указан путь

    Returns:
        int: индекс найденой позиции (если не найдено, то -1)
    """
    ep = try_parse(excel_sheet_str)
    if ep.is_none():
        raise KeyError("Неверный путь к Excel ячейке")
    sheet = wb.get_sheet_by_name(ep.sheet)
    for i, row in enumerate(sheet.iter_rows()):
        if i < pos - 1:
            continue
        cell = row[column_index_from_string(ep.column) - 1]
        try:
            if cell.value == key or key in cell.value:
                return cell.row
        except:
            pass
    return -1

def set_excel_command(wb, key: str, excel_sheet_str: str):
    """функция для добавления функционала в документ

    Args:
        wb (workbook): книга excel
        key (str): ключ для поиска
        excel_sheet_str (str): путь для поиска
    """
    global_index = find_index(wb, excel_sheet_str, key, -1)
    
    @wrapper_error
    @cached
    def get(key: str):
        """функция для получения значения по ключу

        Args:
            key (str): путь

        Returns:
            optional: значение по пути
        """
        index = global_index
        ep = try_parse(key, index)
        if not ep.is_none():
            sheet = wb.get_sheet_by_name(ep.sheet)
            i = ep.column + ep.row
            if sheet[i].value is not None:
                v = sheet[i].value
                return v
        return ""
            
    jin.globals.update(get=get)
    
    @wrapper_error
    @cached
    def find(sheetName: str, what: str = None, pos: int = -1):
        """функция для поиска строки

        Args:
            sheetName (str): путь поиска
            what (str, optional): что искать. Если не указан, то ищеться по первичному ключу из списка.
            pos (int, optional): позиция, с которой необходимо искать. Defaults to -1.

        Returns:
            _type_: _description_
        """
        if what == None:
            return find_index(wb, sheetName, key, pos)
        else:
            return find_index(wb, sheetName, what, pos)
    
    jin.globals.update(find=find)

def print_docx(key: str, excel: str, fromObj: str, template: str, result: str, context: dict = None):
    """функция для формирования документа из шаблона

    Args:
        key (str): ключ, по которому будут искать строку
        excel (str): путь к excel файлу
        fromObj (str): путь для поиска ключа
        template (str): путь к шаблону
        result (str): путь по которому будет сохранен файл
        context (dict, optional): контекст для jinja2. Defaults to None.

    Raises:
        ValueError: неверно указан путь
        magic2functions.jinja2_render_traceback: ошибка jinja2
    """
    doc = DocxTemplate(template)
    jin.globals.update(doc=doc)

    wb = load_workbook(excel, read_only=True, data_only=True)
    set_excel_command(wb, key, fromObj)
    
    try:
        vars = doc.get_undeclared_template_variables(jinja_env=jin)
        if vars:
            s = []
            for var in vars:
                if var not in context:
                    s.append("Определите значение для: " + var)
            if s:
                raise ValueError(", \n".join(s))
        doc.render(context if context != None else {}, jinja_env=jin)
    except:
        raise magic2functions.jinja2_render_traceback(doc.source)
    if "DeleteTable" in context and context["DeleteTable"] == True:
        docx = doc.docx
        if len(docx.tables) > 0:
            docx.tables[0]._element.getparent().remove(docx.tables[0]._element)
        
    doc.save(result)
    wb.close()


def _iter_cols(self, min_col=None, max_col=None, min_row=None,
               max_row=None, values_only=False):
    """итерирование по колонке

    Args:
        min_col (int, optional): минимальная позиция колонки. Defaults to None.
        max_col (int, optional): максимальная позиция колонки. Defaults to None.
        min_row (int, optional): минимальная позиция строки. Defaults to None.
        max_row (int, optional): максимальная позиция строки. Defaults to None.
        values_only (bool, optional): получение только значений. Defaults to False.

    Yields:
        _type_: _description_
    """
    yield from zip(*self.iter_rows(
        min_row=min_row, max_row=max_row,
        min_col=min_col, max_col=max_col, values_only=values_only))

import types

def get_list(excel: str, fromObj: str):
    """функция для получения списка из excel

    Args:
        excel (str): путь к excel файлу
        fromObj (str): путь поиска колонки

    Returns:
        list[str]: список из колонки fromObj по пути excel
    """
    try:
        ep = try_parse(fromObj)
        if ep.is_none():
            return []
        try:
            wb = load_workbook(excel, read_only=True, data_only=True)
        except:
            return []
        sheet = wb.get_sheet_by_name(ep.sheet)
        sheet.iter_cols = types.MethodType(_iter_cols, sheet)
        l =  [ row.value for row in sheet[ep.column] if isinstance(row.value, str) ] 
        return l
    except BaseException as ex:
        logging.error("BaseException",exc_info=True)
        return []